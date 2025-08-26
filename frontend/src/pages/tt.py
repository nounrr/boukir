import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import re
from typing import List, Dict, Optional, Tuple
import unicodedata
from dataclasses import dataclass
import shutil

try:
    from openpyxl import load_workbook  # for reading cell fill colors
except Exception:
    load_workbook = None


@dataclass
class CruiseMatch:
    """Represents a cruise with its Excel row and found manifests"""
    excel_row: int
    cruise_number: str
    cruise_name: str
    manifests: List[str]
    excel_data: Dict


class CruiseDetectorGUI:
    def _init_(self):
        self.root = tk.Tk()
        self.root.title("Détecteur de Manifestes - Croisières")
        self.root.geometry("900x750")

        # Variables
        self.cruise_list_path = tk.StringVar()
        self.manifests_dir_path = tk.StringVar()
        self.cruise_df = None
        self.matches = []
        # Ignore rows with green background in N
        self.ignore_green_var = tk.BooleanVar(value=True)
        self.ignored_row_idxs = set()  # indexes in DataFrame to ignore

        # Post-detection state
        self.last_manifests_dir = None
        self.matched_files = []
        self.ignored_cruise_numbers = []
        self.ignored_files = []
        self.unmatched_pdfs = []
        self.all_pdfs = []
        # Excel classification: summary-only files, detailed-only files, and mixed files
        self.summary_excel_files = []
        self.detailed_excel_files = []
        self.mixed_excel_files = {}
        # Option to include all PDFs for separation (default to True since actions are automatic)
        self.include_all_pdfs_var = tk.BooleanVar(value=True)

        self.setup_ui()
        
    # (removed corrupted duplicate setup_ui)
    def setup_ui(self):
        """Setup the user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)

        # Step 1: Upload cruise list
        step1_frame = ttk.LabelFrame(main_frame, text="Étape 1: Liste des Croisières", padding="5")
        step1_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        step1_frame.columnconfigure(1, weight=1)

        ttk.Label(step1_frame, text="Fichier Excel:").grid(row=0, column=0, sticky=tk.W, padx=5)
        ttk.Entry(step1_frame, textvariable=self.cruise_list_path, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(step1_frame, text="Parcourir...", command=self.browse_cruise_list).grid(row=0, column=2, padx=5)
        ttk.Button(step1_frame, text="Charger", command=self.load_cruise_list).grid(row=0, column=3, padx=5)

        # Step 2: Select manifests directory
        step2_frame = ttk.LabelFrame(main_frame, text="Étape 2: Dossier des Manifestes", padding="5")
        step2_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        step2_frame.columnconfigure(1, weight=1)

        ttk.Label(step2_frame, text="Dossier:").grid(row=0, column=0, sticky=tk.W, padx=5)
        ttk.Entry(step2_frame, textvariable=self.manifests_dir_path, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(step2_frame, text="Parcourir...", command=self.browse_manifests_dir).grid(row=0, column=2, padx=5)
        # Dashboard picker
        ttk.Label(step2_frame, text="Dashboard:").grid(row=1, column=0, sticky=tk.W, padx=5)
        self.dashboard_path = tk.StringVar()
        ttk.Entry(step2_frame, textvariable=self.dashboard_path, width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(step2_frame, text="Parcourir...", command=self.browse_dashboard).grid(row=1, column=2, padx=5)

        # Step 3: Configuration
        step3_frame = ttk.LabelFrame(main_frame, text="Étape 3: Configuration", padding="5")
        step3_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Label(step3_frame, text="Colonne N (numéro):").grid(row=0, column=0, sticky=tk.W, padx=5)
        self.number_column = tk.StringVar(value="N")
        ttk.Entry(step3_frame, textvariable=self.number_column, width=20).grid(row=0, column=1, sticky=tk.W, padx=5)

        ttk.Label(step3_frame, text="Pattern manifeste:").grid(row=0, column=2, sticky=tk.W, padx=10)
        self.pattern_info = ttk.Label(step3_frame, text="Numéro + nom (ex: ABCD-Nom.xlsx)")
        self.pattern_info.grid(row=0, column=3, sticky=tk.W, padx=5)

        # Ignore green background option
        self.chk_ignore_green = ttk.Checkbutton(
            step3_frame,
            text="Ignorer N avec fond vert (00B050)",
            variable=self.ignore_green_var,
        )
        self.chk_ignore_green.grid(row=1, column=0, columnspan=4, sticky=tk.W, padx=5, pady=2)

        # Step 4: Detection and merge buttons
        self.detect_button = ttk.Button(
            main_frame,
            text="🔍 Détecter les Manifestes",
            command=self.detect_manifests,
            state="disabled",
        )
        self.detect_button.grid(row=3, column=0, pady=10, sticky=tk.W)
        self.merge_button = ttk.Button(
            main_frame,
            text="🔗 Merger (dashboard)",
            command=self.merge_to_dashboard,
            state="normal",
        )
        self.merge_button.grid(row=3, column=1, pady=10, sticky=tk.E)

        # Post-detection action buttons (hidden; created but not gridded)
        actions_frame = ttk.Frame(main_frame)
        self.btn_move_processed = ttk.Button(
            actions_frame,
            text="📦 Déplacer 'déjà traités' (0)",
            command=self.move_ignored_manifests,
            state="disabled",
        )
        self.btn_move_pdfs = ttk.Button(
            actions_frame,
            text="🗂 Séparer PDF non traités (0)",
            command=self.move_unmatched_pdfs,
            state="disabled",
        )
        self.chk_all_pdfs = ttk.Checkbutton(
            actions_frame,
            text="Inclure tous les PDF",
            variable=self.include_all_pdfs_var,
            command=self.update_pdf_action_button,
        )
        self.btn_move_summary_excel = ttk.Button(
            actions_frame,
            text="📊 Séparer Excel résumés (0)",
            command=self.move_summary_excel_files,
            state="disabled",
        )
        self.btn_move_correct_excel = ttk.Button(
            actions_frame,
            text="📁 Déplacer Excel corrects (0)",
            command=self.move_correct_excel_files,
            state="disabled",
        )

        # Results area
        results_frame = ttk.LabelFrame(main_frame, text="Résultats", padding="5")
        results_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(4, weight=1)

        columns = ("Ligne", "Numéro", "Nom", "Manifestes", "Statut")
        self.tree = ttk.Treeview(results_frame, columns=columns, show="headings", height=15)
        self.tree.heading("Ligne", text="Ligne Excel")
        self.tree.heading("Numéro", text="N° Croisière")
        self.tree.heading("Nom", text="Nom Croisière")
        self.tree.heading("Manifestes", text="Manifestes Trouvés")
        self.tree.heading("Statut", text="Statut")
        self.tree.column("Ligne", width=80)
        self.tree.column("Numéro", width=100)
        self.tree.column("Nom", width=150)
        self.tree.column("Manifestes", width=300)
        self.tree.column("Statut", width=100)

        v_scrollbar = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Status bar
        self.status_var = tk.StringVar(value="Prêt - Sélectionnez la liste des croisières")
        ttk.Label(main_frame, textvariable=self.status_var, relief="sunken").grid(
            row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5
        )

    def browse_cruise_list(self):
        """Browse for cruise list Excel file"""
        filename = filedialog.askopenfilename(
            title="Sélectionner la liste des croisières",
            filetypes=[("Excel files", ".xlsx *.xls"), ("All files", ".*")]
        )
        if filename:
            self.cruise_list_path.set(filename)

    def browse_manifests_dir(self):
        """Browse for manifests directory"""
        dirname = filedialog.askdirectory(title="Sélectionner le dossier des manifestes")
        if dirname:
            self.manifests_dir_path.set(dirname)

    def browse_dashboard(self):
        """Browse for dashboard Excel file to merge into/output alongside."""
        filename = filedialog.askopenfilename(
            title="Sélectionner le fichier dashboard",
            filetypes=[("Excel files", ".xlsx *.xls"), ("All files", ".*")]
        )
        if filename:
            self.dashboard_path.set(filename)

    def load_cruise_list(self):
        """Load and preview cruise list"""
        try:
            path = self.cruise_list_path.get()
            if not path:
                messagebox.showerror("Erreur", "Veuillez sélectionner un fichier Excel")
                return

            self.status_var.set("Chargement en cours...")
            self.root.update()

            self.cruise_df = pd.read_excel(path)
            num_rows = len(self.cruise_df)
            cols = list(self.cruise_df.columns)

            # Check if N column exists
            n_col = self.number_column.get()
            if n_col not in cols:
                available_cols = ", ".join(cols)
                messagebox.showwarning("Attention",
                                       f"Colonne '{n_col}' non trouvée.\n"
                                       f"Colonnes disponibles: {available_cols}")

            # Compute ignored rows by color if requested and possible
            self.ignored_row_idxs = set()
            if self.ignore_green_var.get():
                try:
                    self.ignored_row_idxs = self._compute_ignored_rows_by_color(Path(path), n_col)
                except Exception as e:
                    messagebox.showwarning("Info", f"Ignorer par couleur non appliqué: {e}")
                    self.ignored_row_idxs = set()

            ignored_info = f" | Ignorés (fond vert): {len(self.ignored_row_idxs)}" if self.ignored_row_idxs else ""
            self.status_var.set(f"✅ Liste chargée: {num_rows} croisières - Colonnes: {', '.join(cols[:5])}...{ignored_info}")
            self.detect_button.config(state="normal")
            messagebox.showinfo(
                "Succès",
                f"Liste chargée avec succès!\n{num_rows} lignes trouvées.\n\n"
                + (f"{len(self.ignored_row_idxs)} ligne(s) seront ignorées (fond vert 00B050).\n\n" if self.ignored_row_idxs else "")
                + "Vous pouvez maintenant sélectionner le dossier des manifestes."
            )

        except Exception as e:
            self.cruise_df = None
            self.detect_button.config(state="disabled")
            self.status_var.set("❌ Erreur lors du chargement")
            messagebox.showerror("Erreur", f"Impossible de charger le fichier:\n{str(e)}")

    def _compute_ignored_rows_by_color(self, excel_path: Path, number_header: str):
        """Return a set of DataFrame indexes to ignore where the N cell has fill color 00B050.
        Only supported for .xlsx files using openpyxl. For other formats, returns empty set.
        """
        ignored = set()
        if excel_path.suffix.lower() != ".xlsx":
            raise RuntimeError("Ignorer par couleur supporté uniquement pour .xlsx")
        if load_workbook is None:
            raise RuntimeError("openpyxl n'est pas disponible")

        def _get_hex_color(cell) -> Optional[str]:
            try:
                fill = getattr(cell, 'fill', None)
                if not fill:
                    return None
                # Prefer start_color, fall back fgColor
                for attr in ('start_color', 'fgColor'):
                    col = getattr(fill, attr, None)
                    if not col:
                        continue
                    # openpyxl Color: try .rgb if it's a string
                    rgb = getattr(col, 'rgb', None)
                    if isinstance(rgb, str) and rgb:
                        return rgb
                    # Some versions might expose .value
                    val = getattr(col, 'value', None)
                    if isinstance(val, str) and val:
                        return val
                    # If a theme/indexed color, we cannot reliably compare to 00B050
                return None
            except Exception:
                return None

        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        # Find header row (assume first row) and locate the column index for number_header
        header_cells = [c.value for c in ws[1]]
        col_idx = None
        for idx, val in enumerate(header_cells, start=1):
            if str(val).strip() == str(number_header).strip():
                col_idx = idx
                break
        if col_idx is None:
            return ignored
        # Iterate rows starting from 2 and collect indexes where fill is green
        GREEN_HEXES = {"00B050", "FF00B050"}
        for r in range(2, ws.max_row + 1):
            c = ws.cell(row=r, column=col_idx)
            rgb = _get_hex_color(c)
            if not rgb:
                continue
            code = str(rgb).upper()
            # Normalize possible ARGB
            code_short = code[2:] if code.startswith('FF') and len(code) == 8 else code
            if code in GREEN_HEXES or code_short in GREEN_HEXES:
                ignored.add(r - 2)  # DataFrame index corresponding to Excel row
        return ignored

    def normalize_cruise_number(self, raw) -> str:
        """Normalize cruise number for pattern matching.
        Handles numeric Excel cells like 1.0 -> '1', 2.0 -> '2', and strings.
        """
        # None/NaN
        try:
            if raw is None or (hasattr(pd, 'isna') and pd.isna(raw)):
                return ""
        except Exception:
            if raw is None:
                return ""

        # Numeric types from Excel
        if isinstance(raw, int):
            return str(raw)
        if isinstance(raw, float):
            try:
                if float(raw).is_integer():
                    return str(int(raw))
                # If not integer, keep as trimmed string without trailing .0
                return str(int(raw))
            except Exception:
                return str(raw).strip()

        # Strings and others
        s = str(raw).strip()
        if not s:
            return ""

        # If contains any letter, treat as alphanumeric code
        if re.search(r"[A-Za-z]", s):
            s = s.upper()
            # Keep only A-Z, 0-9, underscore and hyphen
            s = re.sub(r"[^A-Z0-9_-]", "", s)
            return s

        # Try to parse numeric strings like '1.0', '2,0', '  10 ' safely
        s_compact = s.replace(" ", "")
        # If looks like a simple number with optional single decimal separator
        if re.fullmatch(r"\d+(?:[.,]\d+)?", s_compact):
            try:
                val = float(s_compact.replace(",", "."))
                return str(int(val))
            except Exception:
                pass

        # Fallback: keep digits only but avoid the 1.0 -> 10 issue
        # We'll first remove thousands separators (spaces) and then decimals properly
        s_digits = re.sub(r"\D", "", s_compact)
        if not s_digits:
            return ""
        try:
            return str(int(s_digits))
        except Exception:
            return s_digits

    def find_manifests_for_cruise(self, cruise_number: str, manifests_dir: Path) -> List[str]:
        """Find manifest files matching the cruise number pattern.
        Matches files starting with the cruise_number (case-insensitive),
        optionally followed by -, _ or space, or ending right before extension.
        For numeric cruise numbers, also tries zero-padded variants (2-4 width).
        """
        if not cruise_number:
            return []

        # Build candidate prefixes (case-insensitive comparison)
        prefixes = [cruise_number]
        if cruise_number.isdigit():
            for width in (2, 3, 4):
                p = cruise_number.zfill(width)
                if p not in prefixes:
                    prefixes.append(p)

        exts = {".xlsx", ".xls", ".pdf"}
        results: List[str] = []
        try:
            for fp in manifests_dir.iterdir():
                if not fp.is_file() or fp.suffix.lower() not in exts:
                    continue
                name = fp.name
                name_upper = name.upper()
                for pref in prefixes:
                    pref_u = pref.upper()
                    if name_upper.startswith(pref_u + ".") or \
                       name_upper.startswith(pref_u + "-") or \
                       name_upper.startswith(pref_u + "_") or \
                       name_upper.startswith(pref_u + " ") or \
                       name_upper == pref_u + fp.suffix.upper():
                        results.append(name)
                        break
        except Exception as e:
            print(f"Erreur lors de la recherche dans {manifests_dir}: {e}")

        # Deduplicate and sort
        seen = set()
        ordered: List[str] = []
        for n in results:
            if n not in seen:
                seen.add(n)
                ordered.append(n)
        return sorted(ordered)
        
    def detect_manifests(self):
        """Detect manifests for all cruises"""
        try:
            # Validate inputs
            if self.cruise_df is None or len(self.cruise_df) == 0:
                messagebox.showerror("Erreur", "Veuillez d'abord charger la liste des croisières.\nUtilisez le bouton 'Charger' après avoir sélectionné le fichier Excel.")
                return
                
            manifests_dir = self.manifests_dir_path.get()
            if not manifests_dir:
                messagebox.showerror("Erreur", "Veuillez sélectionner le dossier des manifestes")
                return
                
            manifests_path = Path(manifests_dir)
            if not manifests_path.exists():
                messagebox.showerror("Erreur", "Le dossier des manifestes n'existe pas")
                return
            self.last_manifests_dir = manifests_path
                
            n_col = self.number_column.get()
            if n_col not in self.cruise_df.columns:
                messagebox.showerror("Erreur", f"Colonne '{n_col}' non trouvée dans le fichier Excel")
                return
                
            # Clear previous results
            for item in self.tree.get_children():
                self.tree.delete(item)
                
            self.matches = []
            total_manifests = 0
            ignored_count = 0
            # Reset post-detection sets
            self.matched_files = []
            self.ignored_cruise_numbers = []
            self.ignored_files = []
            self.unmatched_pdfs = []
            self.summary_excel_files = []
            self.detailed_excel_files = []
            
            # Process each cruise
            for index, row in self.cruise_df.iterrows():
                # Skip/flag ignored by color
                if index in self.ignored_row_idxs:
                    ignored_count += 1
                    raw_number = str(row.get(n_col, "")).strip()
                    cruise_number_norm = self.normalize_cruise_number(raw_number)
                    cruise_name = str(row.get("Nom", row.get("Name", row.get("nom", ""))))
                    self.tree.insert("", "end", values=(
                        index + 2,
                        raw_number or "(vide)",
                        cruise_name or "(pas de nom)",
                        "Ignoré",
                        "🟩 Ignoré (00B050)"
                    ))
                    if cruise_number_norm:
                        self.ignored_cruise_numbers.append(cruise_number_norm)
                        # Collect files belonging to ignored cruises
                        for fn in self.find_manifests_for_cruise(cruise_number_norm, manifests_path):
                            self.ignored_files.append(manifests_path / fn)
                    continue
                
                cruise_number = self.normalize_cruise_number(row.get(n_col, ""))
                cruise_name = str(row.get("Nom", row.get("Name", row.get("nom", ""))))
                
                # Find manifests
                manifests = self.find_manifests_for_cruise(cruise_number, manifests_path)
                # Add matched absolute paths
                for fn in manifests:
                    self.matched_files.append(manifests_path / fn)
                total_manifests += len(manifests)
                
                # Determine status
                if manifests:
                    status = f"✅ {len(manifests)} trouvé(s)"
                elif cruise_number:
                    status = "❌ Aucun"
                else:
                    status = "⚠ N° vide"
                
                # Create match object
                match = CruiseMatch(
                    excel_row=index + 2,  # Excel row (1-indexed + header)
                    cruise_number=cruise_number,
                    cruise_name=cruise_name,
                    manifests=manifests,
                    excel_data=row.to_dict()
                )
                self.matches.append(match)
                
                # Add to tree
                manifests_str = ", ".join(manifests) if manifests else "Aucun"
                self.tree.insert("", "end", values=(
                    match.excel_row,
                    cruise_number or "(vide)",
                    cruise_name or "(pas de nom)",
                    manifests_str,
                    status
                ))
                
            # Compute unmatched PDFs in the manifests directory
            exts = {".xlsx", ".xls", ".pdf"}
            all_files = [p for p in self.last_manifests_dir.iterdir() if p.is_file() and p.suffix.lower() in exts]
            matched_set = {p.resolve() for p in self.matched_files}
            ignored_set = {p.resolve() for p in self.ignored_files}
            # Compute all PDFs and unmatched PDFs
            self.all_pdfs = [p for p in all_files if p.suffix.lower() == ".pdf"]
            self.unmatched_pdfs = [
                p for p in self.all_pdfs
                if p.resolve() not in matched_set and p.resolve() not in ignored_set
            ]

            # Update status
            found_count = sum(1 for m in self.matches if m.manifests)
            total_count = len(self.matches) + ignored_count
            
            # Classify Excel files as detailed or summary
            self.classify_excel_files()
            
            extra = f" | Ignorés: {ignored_count}" if ignored_count else ""
            self.status_var.set(f"Détection terminée: {found_count}/{total_count} croisières avec manifestes ({total_manifests} fichiers){extra}")

            # Enable post-detection actions with counts
            self.btn_move_processed.config(
                state=("normal" if self.ignored_files else "disabled"),
                text=f"📦 Déplacer 'déjà traités' ({len(self.ignored_files)})"
            )
            self.update_pdf_action_button()
            
            # Update summary Excel button (count includes summary-only files and mixed files)
            resume_count = len(self.summary_excel_files) + len(self.mixed_excel_files)
            self.btn_move_summary_excel.config(
                state=("normal" if resume_count > 0 else "disabled"),
                text=f"📊 Séparer Excel résumés ({resume_count})"
            )

            # Update correct Excel button (detailed-only files)
            correct_count = len(self.detailed_excel_files)
            self.btn_move_correct_excel.config(
                state=("normal" if correct_count > 0 else "disabled"),
                text=f"📁 Déplacer Excel corrects ({correct_count})"
            )
            
            # Auto-process after detection: move 'déjà traités', PDFs, Excel résumés/échoués, and corrects
            auto_results = self._auto_post_detection()

            # Combined summary popup
            combined = (
                f"Résultats:\n"
                f"• Croisières analysées: {total_count}\n"
                f"• Avec manifestes: {found_count}\n"
                f"• Fichiers manifestes: {total_manifests}\n"
                f"• Ignorées (fond vert): {ignored_count}\n"
                f"• PDF à traiter (avant séparation): {auto_results.get('pdf_before', 0)}\n"
                f"\nActions automatiques:\n"
                f"• 'Déjà traités' déplacés: {auto_results.get('processed_moved', 0)} (échecs: {auto_results.get('processed_failed', 0)})\n"
                f"• PDF déplacés: {auto_results.get('pdf_moved', 0)} (mode: {auto_results.get('pdf_mode', 'non traités')}, échecs: {auto_results.get('pdf_failed', 0)})\n"
                f"• Excel résumés déplacés: {auto_results.get('resume_moved_whole', 0)}, copies (mixtes): {auto_results.get('resume_copied_mixed', 0)}, originaux nettoyés: {auto_results.get('resume_modified_original', 0)}, échecs: {auto_results.get('resume_failed', 0)}, vers 'excel echoues': {auto_results.get('resume_moved_to_failed', 0)}\n"
                f"• Excel corrects déplacés: {auto_results.get('correct_moved', 0)} (échecs: {auto_results.get('correct_failed', 0)})\n"
            )
            messagebox.showinfo("Détection et séparation terminées", combined)
                              
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la détection:\n{str(e)}")
    
    def _unique_dest(self, dest_dir: Path, name: str) -> Path:
        base = Path(name)
        stem, suffix = base.stem, base.suffix
        candidate = dest_dir / name
        i = 1
        while candidate.exists():
            candidate = dest_dir / f"{stem} ({i}){suffix}"
            i += 1
        return candidate

    def _refresh_manifests_dir(self) -> Tuple[bool, bool]:
        """Ensure self.last_manifests_dir matches the UI entry.
        Returns (ok, changed). If ok is False, an error was displayed.
        """
        try:
            path_str = (self.manifests_dir_path.get() or "").strip()
            if not path_str:
                messagebox.showerror("Erreur", "Veuillez sélectionner le dossier des manifestes")
                return False, False
            p = Path(path_str)
            if not p.exists():
                messagebox.showerror("Erreur", "Le dossier des manifestes n'existe pas")
                return False, False
            prev = self.last_manifests_dir.resolve() if self.last_manifests_dir else None
            cur = p.resolve()
            changed = (prev != cur)
            self.last_manifests_dir = p
            return True, changed
        except Exception as e:
            messagebox.showerror("Erreur", f"Chemin invalide: {e}")
            return False, False

    def _move_to_failed_folder(self, src: Path) -> Optional[Path]:
        """Déplacer un fichier Excel échoué vers 'excel echoues' et supprimer l'original s'il reste."""
        try:
            if not self.last_manifests_dir:
                return None
            failed_dir = self.last_manifests_dir / "excel echoues"
            failed_dir.mkdir(exist_ok=True)
            dst = self._unique_dest(failed_dir, src.name)
            shutil.move(str(src), str(dst))
            # Vérifier si le fichier d'origine existe toujours → supprimer
            if src.exists():
                try:
                    src.unlink()
                except Exception:
                    pass
            return dst
        except Exception:
            return None


    def move_ignored_manifests(self, silent: bool = False):
        """Move manifests corresponding to ignored (green) cruises into 'deja traite' folder."""
        try:
            ok, changed = self._refresh_manifests_dir()
            if not ok:
                return {"moved": 0, "failed": 0}
            if changed:
                if not silent:
                    messagebox.showinfo("Info", "Le dossier des manifestes a changé. Veuillez relancer la détection pour 'déjà traités'.")
                return {"moved": 0, "failed": 0}
            if not self.ignored_files:
                if not silent:
                    messagebox.showinfo("Info", "Aucun fichier à déplacer pour 'déjà traités'.")
                return {"moved": 0, "failed": 0}
            target = self.last_manifests_dir / "deja traite"
            target.mkdir(exist_ok=True)
            moved, failed = 0, 0
            for p in self.ignored_files:
                try:
                    dst = self._unique_dest(target, p.name)
                    shutil.move(str(p), str(dst))
                    moved += 1
                except Exception:
                    failed += 1
            self.status_var.set(f"Déplacés 'déjà traités': {moved} fichier(s). Échecs: {failed}")
            if not silent:
                messagebox.showinfo("Déplacement terminé", f"'Déjà traités' déplacés: {moved}\nÉchecs: {failed}\nDossier: {target}")
            # Refresh buttons (files are moved now)
            self.btn_move_processed.config(state="disabled", text="📦 Déplacer 'déjà traités' (0)")
            
            # Update all_pdfs list after moving files
            if self.all_pdfs:
                moved_names = {p.name for p in self.ignored_files if p.suffix.lower() == ".pdf"}
                self.all_pdfs = [p for p in self.all_pdfs if p.name not in moved_names]
                # Update the PDF action button
                self.update_pdf_action_button()
            return {"moved": moved, "failed": failed}
        except Exception as e:
            if not silent:
                messagebox.showerror("Erreur", f"Impossible de déplacer: {e}")
            return {"moved": 0, "failed": 1}

    def move_unmatched_pdfs(self, silent: bool = False):
        """Move PDF files into 'pdf' folder.
        If the 'Inclure tous les PDF' option is enabled, move all PDFs; otherwise only unmatched PDFs.
        """
        try:
            ok, changed = self._refresh_manifests_dir()
            if not ok:
                return {"moved": 0, "failed": 0, "mode": ("tous" if self.include_all_pdfs_var.get() else "non traités")}
            if changed:
                # Re-scan PDFs from the new folder
                exts = {".xlsx", ".xls", ".pdf"}
                all_files = [p for p in self.last_manifests_dir.iterdir() if p.is_file() and p.suffix.lower() in exts]
                self.all_pdfs = [p for p in all_files if p.suffix.lower() == ".pdf"]
                # Unmatched requires detection context; enforce 'all PDFs' mode or ask to detect
                if not self.include_all_pdfs_var.get():
                    if not silent:
                        messagebox.showinfo("Info", "Le dossier a changé. Activez 'Inclure tous les PDF' ou relancez la détection pour séparer les non traités.")
                    self.update_pdf_action_button()
                    return {"moved": 0, "failed": 0, "mode": "non traités"}
                # If including all PDFs, set unmatched list equal to all_pdfs so the move works seamlessly
                self.unmatched_pdfs = list(self.all_pdfs)
            pdf_list = self.all_pdfs if self.include_all_pdfs_var.get() else self.unmatched_pdfs
            if not pdf_list:
                if not silent:
                    msg = "Aucun PDF à déplacer." if self.include_all_pdfs_var.get() else "Aucun PDF non traité à déplacer."
                    messagebox.showinfo("Info", msg)
                return {"moved": 0, "failed": 0, "mode": ("tous" if self.include_all_pdfs_var.get() else "non traités")}
            target = self.last_manifests_dir / "pdf"
            target.mkdir(exist_ok=True)
            moved, failed = 0, 0
            for p in pdf_list:
                try:
                    dst = self._unique_dest(target, p.name)
                    shutil.move(str(p), str(dst))
                    moved += 1
                except Exception:
                    failed += 1
            if self.include_all_pdfs_var.get():
                self.status_var.set(f"PDF déplacés: {moved} fichier(s). Échecs: {failed}")
                if not silent:
                    messagebox.showinfo("Déplacement PDF terminé", f"PDF déplacés: {moved}\nÉchecs: {failed}\nDossier: {target}")
                # After moving all PDFs, clear lists
                self.all_pdfs = []
                self.unmatched_pdfs = []
            else:
                self.status_var.set(f"PDF non traités déplacés: {moved} fichier(s). Échecs: {failed}")
                if not silent:
                    messagebox.showinfo("Déplacement PDF terminé", f"PDF non traités déplacés: {moved}\nÉchecs: {failed}\nDossier: {target}")
                # Remove moved files from unmatched list
                moved_names = {p.name for p in pdf_list}
                self.unmatched_pdfs = [p for p in self.unmatched_pdfs if p.name not in moved_names]
            # Refresh button state
            self.update_pdf_action_button()
            return {"moved": moved, "failed": failed, "mode": ("tous" if self.include_all_pdfs_var.get() else "non traités")}
        except Exception as e:
            if not silent:
                messagebox.showerror("Erreur", f"Impossible de déplacer: {e}")
            return {"moved": 0, "failed": 1, "mode": ("tous" if self.include_all_pdfs_var.get() else "non traités")}

    def update_pdf_action_button(self):
        """Update the PDF separation button label and state based on current lists and option."""
        try:
            if self.include_all_pdfs_var.get():
                count = len(self.all_pdfs)
                text = f"🗂 Séparer tous les PDF ({count})"
                state = "normal" if count > 0 else "disabled"
            else:
                count = len(self.unmatched_pdfs)
                text = f"🗂 Séparer PDF non traités ({count})"
                state = "normal" if count > 0 else "disabled"
            self.btn_move_pdfs.config(state=state, text=text)
        except Exception:
            # Fail-safe: disable button on unexpected issues
            self.btn_move_pdfs.config(state="disabled")
    
    def _sheet_indicator_count(self, headers_or_row: List[str]) -> int:
        """Helper: count detailed indicators in a list of header-like strings."""
        detailed_indicators = [
            'last name', 'first name', 'name', 'firstname', 'surname',
            'passport', 'passport #', 'passport number', 'document', 'document type',
            'nationality', 'nationality code', 'nationality 3-letter code',
            'date of birth', 'dob', 'd.o.b', 'gender', 'sex', 'expiry', 'expires',
            'issue date', 'embark', 'debark', 'cabin', 'function'
        ]
        lower_vals = [str(x).strip().lower() for x in headers_or_row if str(x).strip()]
        return sum(1 for indicator in detailed_indicators if any(indicator in v for v in lower_vals))

    def _is_summary_header(self, headers_or_row: List[str]) -> bool:
        """Heuristic: classify as summary if any header contains 'Female' or 'Male'."""
        vals = [str(x).strip().lower() for x in headers_or_row if str(x).strip()]
        if not vals:
            return False
        # If a column header mentions female or male, treat as summary (even if nationality exists)
        if any('female' in v for v in vals) or any('male' in v for v in vals):
            return True
        return False

    def _detect_sheet_type(self, excel_path: Path, sheet_name: str) -> str:
        """Return 'detailed' or 'summary' for a given sheet by heuristics, scanning top rows for headers."""
        try:
            # Try with header=0 first
            df = pd.read_excel(excel_path, sheet_name=sheet_name, nrows=10)
            if not df.empty:
                headers = [str(c).lower() for c in df.columns]
                # Summary pattern takes precedence
                if self._is_summary_header(headers):
                    return 'summary'
                if self._sheet_indicator_count(headers) >= 3:
                    return 'detailed'
            # Scan first 15 rows without header to find a header-like row
            df2 = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=15, dtype=str)
            for _, row in df2.iterrows():
                row_vals = [str(v) for v in row.tolist()]
                if self._is_summary_header(row_vals):
                    return 'summary'
                if self._sheet_indicator_count(row_vals) >= 3:
                    return 'detailed'
            return 'summary'
        except Exception:
            # If unreadable, assume summary (safer to move/review)
            return 'summary'

    def classify_excel_files(self):
        """Classify Excel files: summary-only, detailed-only, or mixed (per sheet)."""
        if not self.last_manifests_dir:
            return

        self.summary_excel_files = []
        self.detailed_excel_files = []
        self.mixed_excel_files = {}

        excel_files = [p for p in self.last_manifests_dir.iterdir()
                       if p.is_file() and p.suffix.lower() in ['.xlsx', '.xls']]

        for excel_path in excel_files:
            try:
                if excel_path.resolve() in {f.resolve() for f in self.ignored_files}:
                    continue

                # List sheet names
                sheet_names: List[str] = []
                try:
                    xls = pd.ExcelFile(excel_path)
                    sheet_names = xls.sheet_names
                except Exception:
                    # If we cannot list sheets, fallback: treat file as summary
                    self.summary_excel_files.append(excel_path)
                    continue

                detailed_sheets: List[str] = []
                summary_sheets: List[str] = []
                for s in sheet_names:
                    t = self._detect_sheet_type(excel_path, s)
                    if t == 'detailed':
                        detailed_sheets.append(s)
                    else:
                        summary_sheets.append(s)

                if summary_sheets and detailed_sheets:
                    self.mixed_excel_files[excel_path] = {
                        'summary': summary_sheets,
                        'detailed': detailed_sheets,
                    }
                elif summary_sheets and not detailed_sheets:
                    self.summary_excel_files.append(excel_path)
                elif detailed_sheets and not summary_sheets:
                    self.detailed_excel_files.append(excel_path)
                else:
                    # No readable sheets -> treat as summary
                    self.summary_excel_files.append(excel_path)
            except Exception:
                self.summary_excel_files.append(excel_path)

    def _write_sheets_to_excel(self, excel_path: Path, out_path: Path, sheet_names: List[str]):
        """Write selected sheets (by name) from excel_path to out_path using pandas."""
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            for name in sheet_names:
                try:
                    df = pd.read_excel(excel_path, sheet_name=name)
                    df.to_excel(writer, sheet_name=name, index=False)
                except Exception:
                    # Skip sheets that fail to read
                    continue

    def _remove_sheets_in_place(self, excel_path: Path, sheets_to_remove: List[str]) -> bool:
        """Remove given sheets from the workbook in place. Returns True on success."""
        if load_workbook is None or excel_path.suffix.lower() != '.xlsx':
            return False
        try:
            wb = load_workbook(excel_path)
            for s in sheets_to_remove:
                if s in wb.sheetnames and len(wb.sheetnames) > 1:
                    ws = wb[s]
                    wb.remove(ws)
            # Ensure at least one sheet remains
            if not wb.sheetnames:
                return False
            wb.save(excel_path)
            return True
        except Exception:
            return False

    def _hide_sheets_in_place(self, excel_path: Path, sheets_to_hide: List[str]) -> bool:
        """Hide given sheets (veryHidden) in the workbook in place. Returns True on success."""
        if load_workbook is None or excel_path.suffix.lower() != '.xlsx':
            return False
        try:
            wb = load_workbook(excel_path)
            changed = False
            for s in sheets_to_hide:
                if s in wb.sheetnames:
                    try:
                        ws = wb[s]
                        # Use veryHidden to keep UI clean; users can't unhide easily by accident
                        ws.sheet_state = 'veryHidden'
                        changed = True
                    except Exception:
                        # Fallback: try 'hidden'
                        try:
                            ws.sheet_state = 'hidden'
                            changed = True
                        except Exception:
                            pass
            if changed:
                wb.save(excel_path)
            return changed
        except Exception:
            return False

    def move_summary_excel_files(self, silent: bool = False):
        """Process Excel files:
        - Files with only summary sheets: move whole file to 'excel resume'
        - Mixed files: duplicate summary sheets to a copy in 'excel resume' and remove them from the original
        """
        try:
            ok, changed = self._refresh_manifests_dir()
            if not ok:
                return {
                    "moved_whole": 0, "copied_mixed": 0, "modified_original": 0,
                    "failed": 0, "moved_to_failed": 0
                }
            if changed:
                # Refresh classification for the new folder
                self.classify_excel_files()

            resume_dir = self.last_manifests_dir / "excel resume"
            resume_dir.mkdir(exist_ok=True)

            moved_whole, copied_mixed, modified_original, failed = 0, 0, 0, 0
            moved_to_failed = 0
            warnings: List[str] = []
            failed_files: List[str] = []
            failed_files_reasons: List[str] = []
            failed_files_sheets: List[str] = []

            # 1) Move summary-only files
            for p in self.summary_excel_files:
                try:
                    dst = self._unique_dest(resume_dir, p.name)
                    shutil.move(str(p), str(dst))
                    moved_whole += 1
                except Exception:
                    failed += 1
                    failed_files.append(p.name)
                    # Try to get the header and sheet breakdown for explanation
                    try:
                        xls = pd.ExcelFile(p)
                        headers = []
                        resume_sheets = []
                        detailed_sheets = []
                        for sheet in xls.sheet_names:
                            df = pd.read_excel(p, sheet_name=sheet, nrows=1)
                            headers += [str(c) for c in df.columns]
                            # Use the same logic as _detect_sheet_type
                            if self._is_summary_header(df.columns):
                                resume_sheets.append(sheet)
                            elif self._sheet_indicator_count(df.columns) >= 3:
                                detailed_sheets.append(sheet)
                        headers_str = ', '.join(headers)
                        failed_files_reasons.append(f"{p.name}: contient {headers_str} dans l'en-tête donc il est résumé")
                        failed_files_sheets.append(f"{p.name}:\n  Feuilles résumés: {', '.join(resume_sheets) if resume_sheets else 'Aucune'}\n  Feuilles détaillées: {', '.join(detailed_sheets) if detailed_sheets else 'Aucune'}")
                    except Exception:
                        failed_files_reasons.append(f"{p.name}: impossible de lire l'en-tête")
                        failed_files_sheets.append(f"{p.name}: impossible de lire les feuilles")
                    # Move the failed file aside
                    try:
                        dst = self._move_to_failed_folder(p)
                        if dst:
                            moved_to_failed += 1
                    except Exception:
                        pass

            # 2) Handle mixed files
            for p, groups in self.mixed_excel_files.items():
                summary_sheets = groups.get('summary', [])
                detailed_sheets = groups.get('detailed', [])
                copy_path = None
                try:
                    # Create resume copy containing only summary sheets
                    # Always write resume copy as .xlsx for compatibility
                    copy_ext = ".xlsx"
                    copy_name = p.stem + "_resume" + copy_ext
                    copy_path = self._unique_dest(resume_dir, copy_name)
                    self._write_sheets_to_excel(p, copy_path, summary_sheets)
                    copied_mixed += 1

                    # Prefer: hide summary sheets so only detailed remain visible
                    ok = self._hide_sheets_in_place(p, summary_sheets)
                    if ok:
                        modified_original += 1
                    else:
                        # Fallback 1: remove summary sheets (destructive)
                        ok2 = self._remove_sheets_in_place(p, summary_sheets)
                        if ok2:
                            modified_original += 1
                        else:
                            # Fallback 2: give up modifying and move original to failed; remove created copy
                            warnings.append(f"Impossible de masquer/supprimer les feuilles résumés: {p.name}. Original conservé.")
                            try:
                                dst = self._move_to_failed_folder(p)
                                if dst:
                                    moved_to_failed += 1
                            except Exception:
                                pass
                            try:
                                if copy_path and Path(copy_path).exists():
                                    Path(copy_path).unlink(missing_ok=True)
                            except Exception:
                                pass
                except Exception:
                    failed += 1
                    failed_files.append(p.name)
                    try:
                        xls = pd.ExcelFile(p)
                        headers = []
                        resume_sheets = []
                        detailed_sheets_list = []
                        for sheet in xls.sheet_names:
                            df = pd.read_excel(p, sheet_name=sheet, nrows=1)
                            headers += [str(c) for c in df.columns]
                            if self._is_summary_header(df.columns):
                                resume_sheets.append(sheet)
                            elif self._sheet_indicator_count(df.columns) >= 3:
                                detailed_sheets_list.append(sheet)
                        headers_str = ', '.join(headers)
                        failed_files_reasons.append(f"{p.name}: contient {headers_str} dans l'en-tête donc il est résumé")
                        failed_files_sheets.append(f"{p.name}:\n  Feuilles résumés: {', '.join(resume_sheets) if resume_sheets else 'Aucune'}\n  Feuilles détaillées: {', '.join(detailed_sheets_list) if detailed_sheets_list else 'Aucune'}")
                    except Exception:
                        failed_files_reasons.append(f"{p.name}: impossible de lire l'en-tête")
                        failed_files_sheets.append(f"{p.name}: impossible de lire les feuilles")
                    warnings.append(f"Échec de traitement mixte: {p.name}")
                    # Move the failed file aside
                    try:
                        dst = self._move_to_failed_folder(p)
                        if dst:
                            moved_to_failed += 1
                    except Exception:
                        pass
                    # Also remove the created resume copy to avoid leaving it among correct files
                    try:
                        if copy_path and Path(copy_path).exists():
                            Path(copy_path).unlink(missing_ok=True)
                    except Exception:
                        pass

            # Refresh classification and button state
            self.classify_excel_files()
            resume_count = len(self.summary_excel_files) + len(self.mixed_excel_files)
            self.btn_move_summary_excel.config(
                state=("normal" if resume_count > 0 else "disabled"),
                text=f"📊 Séparer Excel résumés ({resume_count})"
            )
            # Update correct Excel button after reclassification
            correct_count = len(self.detailed_excel_files)
            self.btn_move_correct_excel.config(
                state=("normal" if correct_count > 0 else "disabled"),
                text=f"📁 Déplacer Excel corrects ({correct_count})"
            )

            # Status and popup
            self.status_var.set(
                f"Excel résumés: déplacés {moved_whole}, copiés (mixtes) {copied_mixed}, originaux modifiés {modified_original}, échecs {failed}"
            )
            details = (
                f"• Fichiers déplacés (résumé seulement): {moved_whole}\n"
                f"• Copies créées (mixtes): {copied_mixed}\n"
                f"• Originaux nettoyés (feuilles supprimées): {modified_original}\n"
                f"• Échecs: {failed}\n"
                f"• Déplacés vers 'excel echoues': {moved_to_failed}"
            )
            if warnings or failed_files:
                details += "\n\nAvertissements:\n"
                if warnings:
                    details += "- " + "\n- ".join(warnings) + "\n"
                if failed_files:
                    details += "- Fichiers échoués: " + ", ".join(failed_files) + "\n"
                if failed_files_reasons:
                    details += "- Raisons:\n  " + "\n  ".join(failed_files_reasons) + "\n"
                if failed_files_sheets:
                    details += "- Détail des feuilles:\n  " + "\n  ".join(failed_files_sheets)
            if not silent:
                messagebox.showinfo("Séparation Excel résumés", details)

            return {
                "moved_whole": moved_whole,
                "copied_mixed": copied_mixed,
                "modified_original": modified_original,
                "failed": failed,
                "moved_to_failed": moved_to_failed,
            }

        except Exception as e:
            if not silent:
                messagebox.showerror("Erreur", f"Impossible de traiter les Excel résumés: {e}")
            return {
                "moved_whole": 0, "copied_mixed": 0, "modified_original": 0,
                "failed": 1, "moved_to_failed": 0
            }
            
    def run(self):
        """Run the GUI"""
        self.root.mainloop()

    def move_correct_excel_files(self, silent: bool = False):
        """Move detailed Excel files into 'excel correct' folder.

        Now also moves MIXED files (those with at least one detailed sheet) so they
        don't remain in the root even if summary sheets couldn't be hidden/removed.
        """
        try:
            ok, changed = self._refresh_manifests_dir()
            if not ok:
                return {"moved": 0, "failed": 0}
            if changed:
                # Refresh classification for the new folder
                self.classify_excel_files()

            # Build the list of files to move:
            # - all detailed-only files
            # - all mixed files (contain detailed sheets too)
            files_to_move = list(self.detailed_excel_files)
            for p in getattr(self, "mixed_excel_files", {}).keys():
                if p not in files_to_move:
                    files_to_move.append(p)

            # Always ensure the target folder exists
            target = self.last_manifests_dir / "excel correct"
            try:
                target.mkdir(exist_ok=True)
            except Exception:
                pass

            if not files_to_move:
                if not silent:
                    messagebox.showinfo("Info", "Aucun fichier Excel correct à déplacer.")
                return {"moved": 0, "failed": 0}

            moved, failed = 0, 0
            for p in files_to_move:
                try:
                    dst = self._unique_dest(target, p.name)
                    shutil.move(str(p), str(dst))
                    # Clean up if an original lingering handle remains
                    if p.exists():
                        try:
                            p.unlink()
                        except Exception:
                            pass
                    moved += 1
                except Exception:
                    failed += 1

            # Reclassify after moving
            self.classify_excel_files()
            # Update button label/state (include mixed count too)
            correct_count = len(self.detailed_excel_files) + len(self.mixed_excel_files)
            self.btn_move_correct_excel.config(
                state=("normal" if correct_count > 0 else "disabled"),
                text=f"📁 Déplacer Excel corrects ({correct_count})"
            )

            self.status_var.set(f"Excel corrects déplacés: {moved}. Échecs: {failed}")
            if not silent:
                messagebox.showinfo(
                    "Déplacement terminé",
                    f"Excel corrects déplacés: {moved}\nÉchecs: {failed}\nDossier: {target}"
                )
            return {"moved": moved, "failed": failed}
        except Exception as e:
            if not silent:
                messagebox.showerror("Erreur", f"Impossible de déplacer les Excel corrects: {e}")
            return {"moved": 0, "failed": 1}


    def _auto_post_detection(self) -> dict:
        """Run automatic separation steps after detection, silently, and return a summary dict."""
        summary: dict = {}
        # Always include all PDFs in automatic mode
        try:
            self.include_all_pdfs_var.set(True)
        except Exception:
            pass
        # 1) Move ignored (déjà traités)
        res_processed = self.move_ignored_manifests(silent=True)
        summary["processed_moved"] = res_processed.get("moved", 0)
        summary["processed_failed"] = res_processed.get("failed", 0)

        # Capture PDFs count before move for reporting
        summary["pdf_before"] = len(self.all_pdfs)

        # 2) Move PDFs (respect user's 'Inclure tous les PDF' option)
        res_pdfs = self.move_unmatched_pdfs(silent=True)
        summary["pdf_moved"] = res_pdfs.get("moved", 0)
        summary["pdf_failed"] = res_pdfs.get("failed", 0)
        summary["pdf_mode"] = res_pdfs.get("mode", "non traités")

        # 3) Separate Excel résumés and move échoués
        res_resume = self.move_summary_excel_files(silent=True)
        summary["resume_moved_whole"] = res_resume.get("moved_whole", 0)
        summary["resume_copied_mixed"] = res_resume.get("copied_mixed", 0)
        summary["resume_modified_original"] = res_resume.get("modified_original", 0)
        summary["resume_failed"] = res_resume.get("failed", 0)
        summary["resume_moved_to_failed"] = res_resume.get("moved_to_failed", 0)

        # 4) Move Excel corrects
        res_correct = self.move_correct_excel_files(silent=True)
        summary["correct_moved"] = res_correct.get("moved", 0)
        summary["correct_failed"] = res_correct.get("failed", 0)

        return summary

    # --------- Dashboard merge helpers ---------
    def _strip_accents(self, s: str) -> str:
        try:
            return ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
        except Exception:
            return s

    def _normalize_header(self, s: str) -> str:
        raw = str(s or "").strip().lower()
        raw = self._strip_accents(raw)
        raw = raw.replace("_", " ")
        raw = re.sub(r"\s+", " ", raw)
        return raw

    def _find_col(self, columns: List[str], candidates: List[str]) -> Optional[str]:
        cols_norm = [self._normalize_header(str(c)) for c in columns]
        cand_norm = [self._normalize_header(c) for c in candidates]
        # Try exact contains
        for idx, c in enumerate(cols_norm):
            for cand in cand_norm:
                if cand and cand in c:
                    return columns[idx]
        # Try token overlap
        for idx, c in enumerate(cols_norm):
            c_tokens = set(c.split())
            for cand in cand_norm:
                cand_tokens = set(cand.split())
                if cand_tokens and cand_tokens.issubset(c_tokens):
                    return columns[idx]
        return None

    def _score_header_row(self, rows: List[List[str]]) -> int:
        # Not used directly; kept for potential extension
        return 0

    def _read_with_best_header(self, excel_path: Path, sheet_name: str) -> pd.DataFrame:
        # Probe first 25 rows to locate a header row
        try:
            probe = pd.read_excel(excel_path, sheet_name=sheet_name, header=None, nrows=25, dtype=str)
        except Exception:
            try:
                return pd.read_excel(excel_path, sheet_name=sheet_name)
            except Exception:
                return pd.DataFrame()
        # Candidates tokens across fields
        tokens = [
            "first", "prenom", "given", "last", "nom", "surname", "name", "passenger",
            "passport", "passeport", "doc", "id", "nationality", "nationalite", "citizenship",
            "date of birth", "date naissance", "dob", "naissance",
            "gender", "sex", "sexe",
            "embark", "embarquement", "arrival", "arrivee",
            "debark", "disembark", "departure", "depart", "sortie"
        ]
        norm_tokens = [self._normalize_header(t) for t in tokens]
        best_row = 0
        best_score = -1
        for i in range(min(25, len(probe))):
            vals = [self._normalize_header(str(x)) for x in probe.iloc[i].tolist()]
            row_text = " ".join(vals)
            score = sum(1 for t in norm_tokens if t in row_text)
            if score > best_score:
                best_score = score
                best_row = i
        try:
            df = pd.read_excel(excel_path, sheet_name=sheet_name, header=best_row)
        except Exception:
            try:
                df = pd.read_excel(excel_path, sheet_name=sheet_name)
            except Exception:
                return pd.DataFrame()
        # Drop fully-empty rows
        if not df.empty:
            df = df.dropna(how='all')
            # Remove repeated header rows (if any)
            df = df[df.apply(lambda r: not any(isinstance(x, str) and self._normalize_header(str(x)) in [self._normalize_header(c) for c in df.columns] for x in r.values), axis=1)]
        return df

    def _detect_column_map(self, columns: List[str]) -> Dict[str, Optional[str]]:
        """Return which source columns were detected for key fields, for diagnostics."""
        cols = list(columns)
        return {
            "FirstName": self._find_col(
                cols,
                [
                    "first name", "firstname", "given name", "prenom", "prénom", "given",
                    "prénom passager", "prenom passager",
                ],
            ),
            "LastName": self._find_col(
                cols,
                ["last name", "surname", "family name", "nom", "family", "nom de famille"],
            ),
            "FullName": self._find_col(
                cols,
                [
                    "full name", "passenger name", "guest name", "nom complet",
                    "nom et prenom", "nom et prénom", "nom et prénoms",
                ],
            ),
            "Passport": self._find_col(
                cols,
                [
                    "passport", "passport #", "passport no", "passeport", "numero passeport",
                    "n passeport", "n° passeport", "no passeport", "document", "doc number",
                    "id number", "passport number", "passport n°",
                ],
            ),
            "Nationality": self._find_col(
                cols,
                ["nationality", "nationality code", "nationalite", "citizenship", "pays", "country"],
            ),
            "DateOfBirth": self._find_col(
                cols,
                [
                    "date of birth", "dob", "birth date", "d.o.b", "date naissance",
                    "date de naissance", "date naiss",
                ],
            ),
            "Gender": self._find_col(cols, ["gender", "sex", "sexe", "genre"]),
            "DateEntree": self._find_col(
                cols,
                [
                    "embark", "embarkation", "arrival", "arrival date", "date arrivee",
                    "date d'arrivee", "date entree", "entry date", "date d'entree", "eta",
                ],
            ),
            "DateSortie": self._find_col(
                cols,
                [
                    "debark", "disembark", "departure", "departure date", "date sortie",
                    "date depart", "exit date", "date de depart", "etd",
                ],
            ),
        }

    def _map_source_to_dashboard(self, df: pd.DataFrame, source_file: Path, sheet_name: str) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame()
        cols = list(df.columns)
        # Candidate fields (avoid generic 'name' alone to prevent matching 'Ship Name')
        col_first = self._find_col(
            cols,
            [
                "first name", "firstname", "given name", "prenom", "prénom", "given",
                "prénom passager", "prenom passager"
            ],
        )
        col_last = self._find_col(
            cols,
            ["last name", "surname", "family name", "nom", "family", "nom de famille"],
        )
        col_full = self._find_col(
            cols,
            [
                "full name", "passenger name", "guest name", "nom complet",
                "nom et prenom", "nom et prénom", "nom et prénoms",
            ],
        )  # fallback
        col_passport = self._find_col(
            cols,
            [
                "passport", "passport #", "passport no", "passeport", "numero passeport",
                "n passeport", "n° passeport", "no passeport", "document", "doc number",
                "id number", "passport number", "passport n°"
            ],
        )
        col_nat = self._find_col(
            cols,
            ["nationality", "nationality code", "nationalite", "citizenship", "pays", "country"],
        )
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import re
from typing import List, Dict, Optional, Tuple
import unicodedata
from dataclasses import dataclass
import shutil

try:
    from openpyxl import load_workbook  # for reading cell fill colors
except Exception:
    load_workbook = None


@dataclass
class CruiseMatch:
    """Represents a cruise with its Excel row and found manifests"""
    excel_row: int
    cruise_number: str
    cruise_name: str
    manifests: List[str]
    excel_data: Dict


class CruiseDetectorGUI:
    def _init_(self):
        self.root = tk.Tk()
        self.root.title("Détecteur de Manifestes - Croisières")
        self.root.geometry("900x750")

        # Variables
        self.cruise_list_path = tk.StringVar()
        self.manifests_dir_path = tk.StringVar()
        self.cruise_df = None
        self.matches = []
        # Ignore rows with green background in N
        self.ignore_green_var = tk.BooleanVar(value=True)
        self.ignored_row_idxs = set()  # indexes in DataFrame to ignore

        # Post-detection state
        self.last_manifests_dir = None
        self.matched_files = []
        self.ignored_cruise_numbers = []
        self.ignored_files = []
        self.unmatched_pdfs = []
        self.all_pdfs = []
        # Excel classification: su